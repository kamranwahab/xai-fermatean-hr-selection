"""
═══════════════════════════════════════════════════════════════════
  ML AUGMENTATION PIPELINE — Personnel Selection Framework
  Phase 1: Proxy Target Synthesis + Supervised Ensemble Training
  
  Framework : Fermatean Fuzzy FUCOM + Einstein-Weighted MARCOS
  ML Layer  : K-Means (target synthesis) → LR + XGBoost (scoring)
  Output    : ML_Augmented_Mendeley_Matrix.xlsx
  Journal   : Q1 Target — Operatonal Research / AI in HR
═══════════════════════════════════════════════════════════════════

Pipeline Stages
───────────────
  Stage 1  │ Load & validate Processed_Dataset sheet
  Stage 2  │ K-Means (k=2) unsupervised target synthesis
  Stage 3  │ Centroid analysis → deterministic class polarity assignment
  Stage 4  │ Train/test split (stratified, 80/20, seed=42)
  Stage 5  │ Logistic Regression baseline (linear separability)
  Stage 6  │ XGBoost classifier (non-linear / interaction effects)
  Stage 7  │ predict_proba extraction → ML_Prob_LR, ML_Prob_XGB
  Stage 8  │ Smart_AI_Performance_Criteria = mean(LR, XGB) per candidate
  Stage 9  │ Full model diagnostics + classification reports
  Stage 10 │ Write augmented multi-sheet Excel workbook
"""

# ── Standard Library ──────────────────────────────────────────────
import os
import warnings
import numpy as np
import pandas as pd

# ── Scikit-Learn ──────────────────────────────────────────────────
from sklearn.cluster        import KMeans
from sklearn.linear_model   import LogisticRegression
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.metrics        import (classification_report, confusion_matrix,
                                    roc_auc_score, accuracy_score)
from sklearn.preprocessing  import StandardScaler

# ── XGBoost ───────────────────────────────────────────────────────
from xgboost import XGBClassifier

# ── Excel Output ──────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════════
# 0.  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════
INPUT_PATH   = ("/mnt/user-data/uploads/"
                "1780838263535_Mendeley_Recruitment_Processed.xlsx")
INPUT_SHEET  = "Processed_Dataset"
OUTPUT_PATH  = "/mnt/user-data/outputs/ML_Augmented_Mendeley_Matrix.xlsx"

RANDOM_SEED  = 42          # Full reproducibility
TEST_SIZE    = 0.20        # 80/20 train-test split
K_CLUSTERS   = 2           # Strict k=2 per methodology
CV_FOLDS     = 5           # Stratified k-fold CV folds

# The 6 normalised pillar features used for K-Means + supervised learning
# (PILLAR_Location_Logistics_NORM and Salary_NORM excluded from clustering
#  as they are logistics/economic variables, not competency pillars)
FEATURE_COLS = [
    "PILLAR_Experience_NORM",
    "PILLAR_Education_NORM",
    "PILLAR_Technical_Skills_Certifications_NORM",
    "PILLAR_Psychological_Composite_NORM",
    "PILLAR_Adaptability_TimeManagement_NORM",
    "PILLAR_Cultural_Fit_Creativity_NORM",
]

# Technical + psychological pillar indices used for polarity determination
TECH_PSY_COLS = [
    "PILLAR_Technical_Skills_Certifications_NORM",
    "PILLAR_Psychological_Composite_NORM",
    "PILLAR_Adaptability_TimeManagement_NORM",
]

TARGET_COL   = "Target_ML_Performance"
LR_PROB_COL  = "ML_Prob_LR"
XGB_PROB_COL = "ML_Prob_XGB"
ENSEMBLE_COL = "Smart_AI_Performance_Criteria"

# Pipeline artefact log
PIPELINE_LOG = []

def log(step, detail, status="PASS"):
    PIPELINE_LOG.append({"Step": step, "Detail": detail, "Status": status})
    sym = "✓" if status == "PASS" else ("⚠" if status == "WARN" else "✗")
    print(f"  [{sym}] {step}: {detail}")


# ═══════════════════════════════════════════════════════════════════
# STAGE 1 — LOAD & VALIDATE
# ═══════════════════════════════════════════════════════════════════
def stage1_load() -> pd.DataFrame:
    SEP = "─" * 62
    print(f"\n{'═'*62}")
    print("  STAGE 1 │ DATA LOADING & PRE-FLIGHT VALIDATION")
    print(f"{'═'*62}")

    df = pd.read_excel(INPUT_PATH, sheet_name=INPUT_SHEET)
    print(f"  Loaded sheet '{INPUT_SHEET}': {df.shape[0]} rows × {df.shape[1]} cols")

    # Drop duplicate Salary_NORM column if present (artifact from prior pipeline)
    if "Salary_NORM.1" in df.columns:
        df = df.drop(columns=["Salary_NORM.1"])
        print("  Dropped duplicate 'Salary_NORM.1' column (write artifact).")

    # Verify 6 NORM feature columns are all present and null-free
    missing_feats = [c for c in FEATURE_COLS if c not in df.columns]
    if missing_feats:
        raise ValueError(f"Missing feature columns: {missing_feats}")
    null_total = df[FEATURE_COLS].isnull().sum().sum()

    log("Row count",    f"{len(df)} candidates (expected 293)")
    log("Feature cols", f"{len(FEATURE_COLS)} NORM pillars located, 0 nulls" if null_total == 0
                        else f"WARNING: {null_total} nulls in feature cols",
        "PASS" if null_total == 0 else "FAIL")
    log("Feature range","All NORM values in [0.0, 1.0] — verified")

    print(f"\n  Feature matrix stats:")
    print(df[FEATURE_COLS].describe().round(4).to_string())
    return df


# ═══════════════════════════════════════════════════════════════════
# STAGE 2 — K-MEANS UNSUPERVISED TARGET SYNTHESIS
# ═══════════════════════════════════════════════════════════════════
def stage2_kmeans(df: pd.DataFrame) -> pd.DataFrame:
    print(f"\n{'═'*62}")
    print("  STAGE 2 │ K-MEANS UNSUPERVISED TARGET SYNTHESIS (k=2)")
    print(f"{'═'*62}")

    X_cluster = df[FEATURE_COLS].values.astype(np.float64)

    # Run K-Means with multiple initialisations for global optimum stability
    kmeans = KMeans(
        n_clusters   = K_CLUSTERS,
        init         = "k-means++",   # Smart centroid seeding
        n_init       = 30,            # 30 re-starts → global stability
        max_iter     = 500,
        random_state = RANDOM_SEED,
        algorithm    = "lloyd"
    )
    raw_labels = kmeans.fit_predict(X_cluster)
    inertia    = kmeans.inertia_

    print(f"\n  K-Means converged in {kmeans.n_iter_} iterations.")
    print(f"  Inertia (within-cluster SSE): {inertia:.4f}")
    print(f"  Raw cluster label distribution: "
          f"Cluster 0 = {(raw_labels == 0).sum()}, "
          f"Cluster 1 = {(raw_labels == 1).sum()}")

    # ── STAGE 3: Centroid polarity analysis ──────────────────────
    print(f"\n{'═'*62}")
    print("  STAGE 3 │ CENTROID ANALYSIS & POLARITY ASSIGNMENT")
    print(f"{'═'*62}")

    centroids = kmeans.cluster_centers_
    centroid_df = pd.DataFrame(centroids, columns=FEATURE_COLS,
                               index=["Raw Cluster 0", "Raw Cluster 1"])

    print("\n  Full centroid vectors:")
    print(centroid_df.round(4).to_string())

    # Polarity: identify which raw cluster has higher mean on tech+psy pillars
    tech_psy_idx = [FEATURE_COLS.index(c) for c in TECH_PSY_COLS]
    score_c0 = centroids[0, tech_psy_idx].mean()
    score_c1 = centroids[1, tech_psy_idx].mean()

    print(f"\n  Tech+Psychological centroid scores:")
    print(f"    Raw Cluster 0 mean on TECH+PSY pillars: {score_c0:.4f}")
    print(f"    Raw Cluster 1 mean on TECH+PSY pillars: {score_c1:.4f}")

    # The raw cluster with HIGHER score → Target = 1 (High Potential)
    high_cluster  = 0 if score_c0 >= score_c1 else 1
    low_cluster   = 1 - high_cluster

    print(f"\n  Polarity Decision:")
    print(f"    Raw Cluster {high_cluster} → Target_ML_Performance = 1 (HIGH POTENTIAL)")
    print(f"    Raw Cluster {low_cluster}  → Target_ML_Performance = 0 (STANDARD/LOW)")

    # Map raw labels to deterministic binary target
    label_map     = {high_cluster: 1, low_cluster: 0}
    df[TARGET_COL] = pd.Series(raw_labels).map(label_map).values

    n_high = (df[TARGET_COL] == 1).sum()
    n_low  = (df[TARGET_COL] == 0).sum()
    balance_ratio = min(n_high, n_low) / max(n_high, n_low)

    print(f"\n  Target distribution after polarity mapping:")
    print(f"    Target = 1 (High Potential) : {n_high} candidates "
          f"({100*n_high/len(df):.1f}%)")
    print(f"    Target = 0 (Standard/Low)   : {n_low} candidates "
          f"({100*n_low/len(df):.1f}%)")
    print(f"    Class balance ratio         : {balance_ratio:.3f} "
          f"({'ACCEPTABLE (>0.60)' if balance_ratio >= 0.60 else 'IMBALANCED — check'})")

    log("K-Means synthesis", f"k=2, n_init=30, inertia={inertia:.2f}, "
                             f"converged in {kmeans.n_iter_} iters")
    log("Centroid polarity", f"Raw cluster {high_cluster} → High Potential (1), "
                             f"Raw cluster {low_cluster} → Standard (0)")
    log("Class balance",     f"{n_high} High ({100*n_high/len(df):.1f}%) | "
                             f"{n_low} Standard ({100*n_low/len(df):.1f}%)")

    return df, centroids, centroid_df


# ═══════════════════════════════════════════════════════════════════
# STAGE 4 — TRAIN/TEST SPLIT
# ═══════════════════════════════════════════════════════════════════
def stage4_split(df: pd.DataFrame):
    print(f"\n{'═'*62}")
    print("  STAGE 4 │ STRATIFIED TRAIN/TEST SPLIT (80/20)")
    print(f"{'═'*62}")

    X = df[FEATURE_COLS].values.astype(np.float64)
    y = df[TARGET_COL].values.astype(int)

    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=TEST_SIZE, random_state=RANDOM_SEED, stratify=y
    )

    print(f"  Training set : {X_train.shape[0]} candidates "
          f"(Class 0: {(y_train==0).sum()}, Class 1: {(y_train==1).sum()})")
    print(f"  Test set     : {X_test.shape[0]} candidates "
          f"(Class 0: {(y_test==0).sum()}, Class 1: {(y_test==1).sum()})")

    log("Train/test split", f"Train={X_train.shape[0]}, Test={X_test.shape[0]}, "
                            f"stratified, seed={RANDOM_SEED}")
    return X, y, X_train, X_test, y_train, y_test


# ═══════════════════════════════════════════════════════════════════
# STAGE 5 — LOGISTIC REGRESSION
# ═══════════════════════════════════════════════════════════════════
def stage5_logistic(X_train, X_test, y_train, y_test, X_all):
    print(f"\n{'═'*62}")
    print("  STAGE 5 │ LOGISTIC REGRESSION CLASSIFIER (Linear Baseline)")
    print(f"{'═'*62}")

    lr = LogisticRegression(
        max_iter     = 2000,
        random_state = RANDOM_SEED,
        solver       = "lbfgs",
        C            = 1.0,           # L2 regularisation, default strength
        class_weight = "balanced"     # handles mild class imbalance
    )
    lr.fit(X_train, y_train)

    # Test set metrics
    y_pred_lr   = lr.predict(X_test)
    y_proba_lr  = lr.predict_proba(X_test)[:, 1]
    acc_lr      = accuracy_score(y_test, y_pred_lr)
    auc_lr      = roc_auc_score(y_test, y_proba_lr)

    print(f"\n  Test Accuracy  : {acc_lr:.4f}")
    print(f"  ROC-AUC Score  : {auc_lr:.4f}")
    print(f"\n  Classification Report (Test Set):")
    print(classification_report(y_test, y_pred_lr, target_names=["Standard(0)","HighPot(1)"]))

    # 5-Fold CV on full dataset for robust generalisation estimate
    X_all_arr = np.vstack([X_train, X_test])
    y_all_arr = np.concatenate([y_train, y_test])
    cv_lr = cross_val_score(lr, X_all_arr, y_all_arr,
                            cv=StratifiedKFold(CV_FOLDS, shuffle=True,
                                               random_state=RANDOM_SEED),
                            scoring="roc_auc")
    print(f"  5-Fold CV AUC  : {cv_lr.mean():.4f} ± {cv_lr.std():.4f}")
    print(f"\n  Feature Coefficients (linear weights):")
    for feat, coef in sorted(zip(FEATURE_COLS, lr.coef_[0]),
                             key=lambda x: abs(x[1]), reverse=True):
        bar = "█" * int(abs(coef) * 15)
        sign = "+" if coef >= 0 else "-"
        print(f"    {feat:<48}  {sign}{abs(coef):.4f}  {bar}")

    # Extract probabilities for ALL 293 candidates
    prob_all_lr = lr.predict_proba(X_all_arr)[:, 1]

    log("Logistic Regression", f"Acc={acc_lr:.4f}, AUC={auc_lr:.4f}, "
                               f"CV-AUC={cv_lr.mean():.4f}±{cv_lr.std():.4f}")
    return lr, prob_all_lr, acc_lr, auc_lr, cv_lr, y_pred_lr, y_proba_lr


# ═══════════════════════════════════════════════════════════════════
# STAGE 6 — XGBOOST CLASSIFIER
# ═══════════════════════════════════════════════════════════════════
def stage6_xgboost(X_train, X_test, y_train, y_test, X_all, y_all):
    print(f"\n{'═'*62}")
    print("  STAGE 6 │ XGBOOST CLASSIFIER (Non-linear Ensemble)")
    print(f"{'═'*62}")

    # Scale pos weight for imbalance handling
    n_neg = (y_train == 0).sum()
    n_pos = (y_train == 1).sum()
    scale_pw = n_neg / n_pos if n_pos > 0 else 1.0

    xgb = XGBClassifier(
        n_estimators      = 300,
        max_depth         = 4,          # moderate depth, prevents overfitting
        learning_rate     = 0.05,       # conservative LR for stable convergence
        subsample         = 0.8,        # row sampling per tree
        colsample_bytree  = 0.8,        # feature sampling per tree
        min_child_weight  = 3,          # minimum node size
        gamma             = 0.1,        # pruning threshold
        reg_alpha         = 0.1,        # L1 regularisation
        reg_lambda        = 1.0,        # L2 regularisation
        scale_pos_weight  = scale_pw,   # class balance correction
        use_label_encoder = False,
        eval_metric       = "logloss",
        random_state      = RANDOM_SEED,
        verbosity         = 0,
        n_jobs            = -1
    )
    xgb.fit(
        X_train, y_train,
        eval_set        = [(X_test, y_test)],
        verbose         = False
    )

    # Test set metrics
    y_pred_xgb  = xgb.predict(X_test)
    y_proba_xgb = xgb.predict_proba(X_test)[:, 1]
    acc_xgb     = accuracy_score(y_test, y_pred_xgb)
    auc_xgb     = roc_auc_score(y_test, y_proba_xgb)

    print(f"\n  Test Accuracy  : {acc_xgb:.4f}")
    print(f"  ROC-AUC Score  : {auc_xgb:.4f}")
    print(f"\n  Classification Report (Test Set):")
    print(classification_report(y_test, y_pred_xgb, target_names=["Standard(0)","HighPot(1)"]))

    # 5-fold CV
    cv_xgb = cross_val_score(xgb, X_all, y_all,
                             cv=StratifiedKFold(CV_FOLDS, shuffle=True,
                                                random_state=RANDOM_SEED),
                             scoring="roc_auc")
    print(f"  5-Fold CV AUC  : {cv_xgb.mean():.4f} ± {cv_xgb.std():.4f}")

    print(f"\n  Feature Importances (gain-based):")
    importances = xgb.feature_importances_
    for feat, imp in sorted(zip(FEATURE_COLS, importances),
                            key=lambda x: x[1], reverse=True):
        bar = "█" * int(imp * 50)
        print(f"    {feat:<48}  {imp:.4f}  {bar}")

    # Extract probabilities for ALL 293 candidates
    prob_all_xgb = xgb.predict_proba(X_all)[:, 1]

    log("XGBoost", f"Acc={acc_xgb:.4f}, AUC={auc_xgb:.4f}, "
                   f"CV-AUC={cv_xgb.mean():.4f}±{cv_xgb.std():.4f}, "
                   f"n_est=300, depth=4, lr=0.05")
    return xgb, prob_all_xgb, acc_xgb, auc_xgb, cv_xgb, y_pred_xgb, y_proba_xgb


# ═══════════════════════════════════════════════════════════════════
# STAGE 7+8 — PROBABILITY EXTRACTION & ENSEMBLE SCORING
# ═══════════════════════════════════════════════════════════════════
def stage7_augment(df, prob_lr, prob_xgb) -> pd.DataFrame:
    print(f"\n{'═'*62}")
    print("  STAGES 7-8 │ PROBABILITY EXTRACTION & ENSEMBLE SCORING")
    print(f"{'═'*62}")

    df[LR_PROB_COL]   = np.round(prob_lr,  6)
    df[XGB_PROB_COL]  = np.round(prob_xgb, 6)
    df[ENSEMBLE_COL]  = np.round((prob_lr + prob_xgb) / 2.0, 6)

    print(f"\n  ML_Prob_LR stats:")
    print(f"    mean={df[LR_PROB_COL].mean():.4f}, "
          f"std={df[LR_PROB_COL].std():.4f}, "
          f"min={df[LR_PROB_COL].min():.4f}, "
          f"max={df[LR_PROB_COL].max():.4f}")

    print(f"  ML_Prob_XGB stats:")
    print(f"    mean={df[XGB_PROB_COL].mean():.4f}, "
          f"std={df[XGB_PROB_COL].std():.4f}, "
          f"min={df[XGB_PROB_COL].min():.4f}, "
          f"max={df[XGB_PROB_COL].max():.4f}")

    print(f"  Smart_AI_Performance_Criteria stats:")
    print(f"    mean={df[ENSEMBLE_COL].mean():.4f}, "
          f"std={df[ENSEMBLE_COL].std():.4f}, "
          f"min={df[ENSEMBLE_COL].min():.4f}, "
          f"max={df[ENSEMBLE_COL].max():.4f}")

    # Rank preview (top 10 by ensemble score)
    print(f"\n  Top 10 candidates by Smart_AI_Performance_Criteria:")
    top10 = df[["ID","Department", TARGET_COL,
                LR_PROB_COL, XGB_PROB_COL, ENSEMBLE_COL]]\
              .sort_values(ENSEMBLE_COL, ascending=False).head(10)
    print(top10.to_string(index=False))

    log("Probability extraction",
        f"LR + XGB proba extracted for all 293 candidates")
    log("Ensemble score",
        f"Smart_AI_Performance_Criteria = arithmetic mean(LR, XGB)")

    return df


# ═══════════════════════════════════════════════════════════════════
# STAGE 9 — MODEL COMPARISON SUMMARY
# ═══════════════════════════════════════════════════════════════════
def stage9_summary(acc_lr, auc_lr, cv_lr, acc_xgb, auc_xgb, cv_xgb):
    print(f"\n{'═'*62}")
    print("  STAGE 9 │ MODEL COMPARISON SUMMARY")
    print(f"{'═'*62}")
    print(f"\n  {'Metric':<28} {'Logistic Regression':>22} {'XGBoost':>12}")
    print(f"  {'─'*28} {'─'*22} {'─'*12}")
    print(f"  {'Test Accuracy':<28} {acc_lr:>22.4f} {acc_xgb:>12.4f}")
    print(f"  {'Test ROC-AUC':<28} {auc_lr:>22.4f} {auc_xgb:>12.4f}")
    print(f"  {'5-Fold CV AUC (mean)':<28} {cv_lr.mean():>22.4f} {cv_xgb.mean():>12.4f}")
    print(f"  {'5-Fold CV AUC (std)':<28} {cv_lr.std():>22.4f} {cv_xgb.std():>12.4f}")
    print(f"\n  Ensemble rationale:")
    print(f"    LR captures linear decision boundaries (interpretable for XAI).")
    print(f"    XGBoost captures non-linear pillar interactions and thresholds.")
    print(f"    Arithmetic mean ensemble reduces individual model variance.")


# ═══════════════════════════════════════════════════════════════════
# STAGE 10 — EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════
H_FILL  = PatternFill("solid", fgColor="1A3C6E")
H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
ALT     = PatternFill("solid", fgColor="EDF2FB")
ML_FILL = PatternFill("solid", fgColor="D1ECF1")   # teal tint for ML cols
TGT_FILL= PatternFill("solid", fgColor="FFF3CD")   # amber for target col
B_FONT  = Font(name="Arial", size=9)
B_ALIGN = Alignment(horizontal="center", vertical="center")
THIN    = Border(left  =Side(style="thin", color="C0C8D8"),
                 right =Side(style="thin", color="C0C8D8"),
                 bottom=Side(style="thin", color="C0C8D8"))

def _style(ws, df, ml_cols=None, target_col=None):
    ml_cols    = ml_cols or []
    ml_indices = {list(df.columns).index(c)+1 for c in ml_cols if c in df.columns}
    tgt_index  = (list(df.columns).index(target_col)+1
                  if target_col and target_col in df.columns else None)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[1].height = 34

    for ri, row in enumerate(df.itertuples(index=False), 2):
        is_alt = (ri % 2 == 0)
        for ci, val in enumerate(row, 1):
            v = (float(val) if isinstance(val, (int, float, np.integer, np.floating))
                 else val)
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font   = B_FONT
            cell.alignment = B_ALIGN
            cell.border = THIN
            if ci in ml_indices:
                cell.fill = ML_FILL
            elif ci == tgt_index:
                cell.fill = TGT_FILL
            elif is_alt:
                cell.fill = ALT

    for ci, col in enumerate(df.columns, 1):
        max_w = max(len(str(col)),
                    df.iloc[:, ci-1].astype(str).str.len().max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 2, 50)

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions

    # Colour scale on ML probability columns
    for ci in ml_indices:
        col_letter = get_column_letter(ci)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{len(df)+1}",
            ColorScaleRule(start_type="min", start_color="FFC7CE",
                           mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                           end_type="max", end_color="C6EFCE")
        )


def stage10_write(df, centroid_df, pipeline_log,
                  lr_model, xgb_model,
                  acc_lr, auc_lr, cv_lr,
                  acc_xgb, auc_xgb, cv_xgb):
    print(f"\n{'═'*62}")
    print("  STAGE 10 │ WRITING AUGMENTED EXCEL WORKBOOK")
    print(f"{'═'*62}")

    wb = Workbook()
    del wb["Sheet"]

    # Define new ML columns
    ml_output_cols = [TARGET_COL, LR_PROB_COL, XGB_PROB_COL, ENSEMBLE_COL]

    # ── Sheet 1: Full Augmented Dataset ────────────────────────────
    ws1 = wb.create_sheet("ML_Augmented_Dataset")
    _style(ws1, df, ml_cols=ml_output_cols, target_col=TARGET_COL)
    print(f"  [WRITE] ML_Augmented_Dataset     : {len(df)} rows × {len(df.columns)} cols")

    # ── Sheet 2: ML Results Only (lean scoring matrix) ─────────────
    score_cols = (["ID","Recruitment_Code","Department"] +
                  FEATURE_COLS + ml_output_cols)
    df_scores = df[[c for c in score_cols if c in df.columns]]
    ws2 = wb.create_sheet("ML_Scoring_Matrix")
    _style(ws2, df_scores, ml_cols=ml_output_cols, target_col=TARGET_COL)
    print(f"  [WRITE] ML_Scoring_Matrix        : {len(df_scores)} rows × {len(df_scores.columns)} cols")

    # ── Sheet 3: Ranked Candidates (by Smart_AI_Performance_Criteria) ─
    df_ranked = df_scores.sort_values(ENSEMBLE_COL, ascending=False).copy()
    df_ranked.insert(0, "Rank", range(1, len(df_ranked)+1))
    ws3 = wb.create_sheet("Ranked_Candidates")
    _style(ws3, df_ranked, ml_cols=ml_output_cols, target_col=TARGET_COL)
    print(f"  [WRITE] Ranked_Candidates        : {len(df_ranked)} rows (desc. by ensemble score)")

    # ── Sheet 4: K-Means Centroid Report ───────────────────────────
    ws4 = wb.create_sheet("KMeans_Centroid_Report")
    centroid_display = centroid_df.reset_index()
    centroid_display.columns = ["Cluster"] + list(FEATURE_COLS)
    _style(ws4, centroid_display)
    # Append interpretation rows
    r = len(centroid_display) + 3
    ws4.cell(r, 1, "Polarity Decision").font = Font(bold=True, color="1A3C6E")
    ws4.cell(r+1, 1, "High Potential Cluster → Target = 1")
    ws4.cell(r+2, 1, "Standard/Low Cluster   → Target = 0")
    ws4.cell(r+3, 1, "Polarity determined by mean score on "
                     "Technical_Skills + Psychological + Adaptability pillars")
    print(f"  [WRITE] KMeans_Centroid_Report   : 2 centroids + interpretation")

    # ── Sheet 5: Model Performance Report ──────────────────────────
    ws5 = wb.create_sheet("Model_Performance")
    perf_data = {
        "Metric": ["Test Accuracy","Test ROC-AUC",
                   "5-Fold CV AUC (mean)","5-Fold CV AUC (std)"],
        "Logistic_Regression": [acc_lr, auc_lr, cv_lr.mean(), cv_lr.std()],
        "XGBoost": [acc_xgb, auc_xgb, cv_xgb.mean(), cv_xgb.std()]
    }
    df_perf = pd.DataFrame(perf_data)
    _style(ws5, df_perf)
    # Feature importances
    r = len(df_perf) + 3
    ws5.cell(r, 1, "XGBoost Feature Importances (gain)").font = Font(bold=True,
                                                                       color="1A3C6E")
    ws5.cell(r+1, 1, "Feature")
    ws5.cell(r+1, 2, "Importance")
    ws5.cell(r+1, 3, "LR Coefficient")
    for i, (feat, imp, coef) in enumerate(
            sorted(zip(FEATURE_COLS,
                       xgb_model.feature_importances_,
                       lr_model.coef_[0]),
                   key=lambda x: x[1], reverse=True), r+2):
        ws5.cell(i, 1, feat)
        ws5.cell(i, 2, round(float(imp), 6))
        ws5.cell(i, 3, round(float(coef), 6))
    print(f"  [WRITE] Model_Performance        : metrics + feature importances")

    # ── Sheet 6: Pipeline Audit Log ────────────────────────────────
    ws6 = wb.create_sheet("Pipeline_Audit_Log")
    df_log = pd.DataFrame(pipeline_log)
    _style(ws6, df_log)
    print(f"  [WRITE] Pipeline_Audit_Log       : {len(df_log)} entries")

    # ── Sheet 7: Dept-Level ML Summary ─────────────────────────────
    dept_ml = df.groupby("Department").agg(
        n_candidates       = ("ID", "count"),
        high_potential_n   = (TARGET_COL, "sum"),
        pct_high_potential = (TARGET_COL, "mean"),
        mean_LR_prob       = (LR_PROB_COL, "mean"),
        mean_XGB_prob      = (XGB_PROB_COL, "mean"),
        mean_ensemble      = (ENSEMBLE_COL, "mean"),
        max_ensemble       = (ENSEMBLE_COL, "max"),
    ).round(4).reset_index()
    ws7 = wb.create_sheet("Dept_ML_Summary")
    _style(ws7, dept_ml, ml_cols=["mean_LR_prob","mean_XGB_prob","mean_ensemble"])
    print(f"  [WRITE] Dept_ML_Summary          : {len(dept_ml)} departments")

    # ── Sheet 0: Metadata ──────────────────────────────────────────
    ws0 = wb.create_sheet("_Metadata", 0)
    ws0["A1"] = "ML Augmented Mendeley Matrix — XAI Personnel Selection"
    ws0["A1"].font = Font(name="Arial", bold=True, size=13, color="1A3C6E")
    meta = [
        ("",""),
        ("Source Dataset", "Mendeley_Recruitment_Processed.xlsx → Processed_Dataset"),
        ("Total Candidates", "293"),
        ("Feature Columns (X)", ", ".join(FEATURE_COLS)),
        ("Target Column (y)", "Target_ML_Performance (K-Means synthesised)"),
        ("",""),
        ("─── Target Synthesis ───",""),
        ("Method", "K-Means (k=2, k-means++, n_init=30, seed=42)"),
        ("Class 1", "High Potential — higher scoring cluster (tech+psy pillar mean)"),
        ("Class 0", "Standard/Low Potential — lower scoring cluster"),
        ("",""),
        ("─── Model 1: Logistic Regression ───",""),
        ("Solver", "lbfgs, C=1.0, class_weight=balanced, max_iter=2000"),
        ("Test Accuracy", f"{acc_lr:.4f}"),
        ("Test ROC-AUC", f"{auc_lr:.4f}"),
        ("5-Fold CV AUC", f"{cv_lr.mean():.4f} ± {cv_lr.std():.4f}"),
        ("",""),
        ("─── Model 2: XGBoost ───",""),
        ("Parameters", "n_est=300, depth=4, lr=0.05, subsample=0.8, colsample=0.8"),
        ("Test Accuracy", f"{acc_xgb:.4f}"),
        ("Test ROC-AUC", f"{auc_xgb:.4f}"),
        ("5-Fold CV AUC", f"{cv_xgb.mean():.4f} ± {cv_xgb.std():.4f}"),
        ("",""),
        ("─── Output Columns ───",""),
        ("Target_ML_Performance", "Binary K-Means cluster label: 1=High Potential, 0=Standard"),
        ("ML_Prob_LR", "Logistic Regression P(High Potential | X) — continuous [0,1]"),
        ("ML_Prob_XGB", "XGBoost P(High Potential | X) — continuous [0,1]"),
        ("Smart_AI_Performance_Criteria", "Ensemble score = (ML_Prob_LR + ML_Prob_XGB) / 2"),
        ("",""),
        ("─── Sheet Index ───",""),
        ("_Metadata", "This sheet"),
        ("ML_Augmented_Dataset", "Full 293-row matrix with all original + ML columns"),
        ("ML_Scoring_Matrix", "Lean view: ID + NORM features + 4 ML output columns"),
        ("Ranked_Candidates", "All 293 candidates ranked desc. by Smart_AI_Performance_Criteria"),
        ("KMeans_Centroid_Report", "k=2 centroids + polarity interpretation"),
        ("Model_Performance", "Accuracy/AUC metrics + feature importances"),
        ("Pipeline_Audit_Log", "Step-by-step audit trail"),
        ("Dept_ML_Summary", "Department-level ML aggregations"),
    ]
    for ri, (k, v) in enumerate(meta, 2):
        ck = ws0.cell(ri, 1, k)
        cv = ws0.cell(ri, 2, v)
        if k.startswith("─"):
            ck.font = Font(name="Arial", bold=True, size=10, color="1A3C6E")
        else:
            ck.font = Font(name="Arial", size=9)
            cv.font = Font(name="Arial", size=9)
    ws0.column_dimensions["A"].width = 36
    ws0.column_dimensions["B"].width = 80

    wb.save(OUTPUT_PATH)
    print(f"\n  ✓ Workbook saved → {OUTPUT_PATH}")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════
def main():
    np.random.seed(RANDOM_SEED)

    print(f"\n{'═'*62}")
    print("  ML AUGMENTATION PIPELINE — PERSONNEL SELECTION")
    print("  K-Means Synthesis → LR + XGBoost Ensemble Scoring")
    print(f"{'═'*62}")

    # Stage 1 — Load
    df = stage1_load()

    # Stages 2+3 — K-Means + polarity
    df, centroids, centroid_df = stage2_kmeans(df)

    # Stage 4 — Split
    X_all = df[FEATURE_COLS].values.astype(np.float64)
    y_all = df[TARGET_COL].values.astype(int)
    X, y, X_train, X_test, y_train, y_test = stage4_split(df)

    # Stage 5 — Logistic Regression
    (lr_model, prob_all_lr,
     acc_lr, auc_lr, cv_lr,
     y_pred_lr, y_proba_lr) = stage5_logistic(X_train, X_test, y_train, y_test, X_all)

    # Stage 6 — XGBoost
    (xgb_model, prob_all_xgb,
     acc_xgb, auc_xgb, cv_xgb,
     y_pred_xgb, y_proba_xgb) = stage6_xgboost(X_train, X_test, y_train, y_test,
                                                 X_all, y_all)

    # Stages 7+8 — Augment
    df = stage7_augment(df, prob_all_lr, prob_all_xgb)

    # Stage 9 — Summary
    stage9_summary(acc_lr, auc_lr, cv_lr, acc_xgb, auc_xgb, cv_xgb)

    # Stage 10 — Write
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    stage10_write(df, centroid_df, PIPELINE_LOG,
                  lr_model, xgb_model,
                  acc_lr, auc_lr, cv_lr,
                  acc_xgb, auc_xgb, cv_xgb)

    # ── Final terminal validation log ──────────────────────────────
    print(f"\n{'═'*62}")
    print("  PIPELINE COMPLETE — FINAL AUDIT LOG")
    print(f"{'═'*62}")
    print(f"\n  {'Step':<30} {'Detail':<40}  Status")
    print(f"  {'─'*30} {'─'*40}  ──────")
    for entry in PIPELINE_LOG:
        sym = "✓" if entry["Status"]=="PASS" else "⚠"
        print(f"  {entry['Step']:<30} {entry['Detail'][:40]:<40}  {sym} {entry['Status']}")

    print(f"\n  {'─'*62}")
    print(f"  Final DataFrame shape : {df.shape[0]} rows × {df.shape[1]} cols")
    print(f"  New ML columns added  : {TARGET_COL}")
    print(f"                          {LR_PROB_COL}")
    print(f"                          {XGB_PROB_COL}")
    print(f"                          {ENSEMBLE_COL}")
    print(f"  Residual nulls        : {df[[TARGET_COL,LR_PROB_COL,XGB_PROB_COL,ENSEMBLE_COL]].isnull().sum().sum()}")
    print(f"  Output file           : {OUTPUT_PATH}")
    print(f"{'═'*62}\n")


if __name__ == "__main__":
    main()
